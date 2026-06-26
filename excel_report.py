"""
Generates the paid deliverable Excel workbook:
  Sheet 1: Mismatch Summary  — every flagged issue, severity, the numbers, what to do
  Sheet 2: Document Checklist — what to keep ready, checkbox-style tracker
No invented data — every row traces back to a flag from analysis.py or a
fixed checklist entry tied to that flag's bucket. Nothing here should ever
show a number that didn't come from the user's own confirmed Form16/AIS figures.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SEVERITY_FILL = {
    "red": PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid"),
    "amber": PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid"),
    "green": PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid"),
}

BUCKET_LABELS = {
    "salary_mismatch": "Salary mismatch (Form16 vs AIS)",
    "interest_income_omitted": "Interest income not declared",
    "tds_mismatch": "TDS credit mismatch",
    "capital_gains_mismatch": "Capital gains discrepancy",
    "high_value_vs_income": "High-value transactions vs declared income",
    "gst_itr_gap": "GST turnover vs ITR turnover gap",
    "exemption_swap": "Exemption switched between original and revised ITR",
    "stale_ais": "AIS data may be outdated",
}

CHECKLIST_BY_BUCKET = {
    "salary_mismatch": ["Corrected Form16 from employer (if AIS figure is wrong)",
                        "Salary slips for the financial year",
                        "Bank statement showing salary credits"],
    "interest_income_omitted": ["Bank interest certificate (savings/FD/RD)",
                                 "Updated/revised ITR including this interest"],
    "tds_mismatch": ["Form 26AS download",
                      "Corrected Form16 or TDS certificate from deductor"],
    "capital_gains_mismatch": ["Broker/AMC capital gains statement",
                                "Sale deed or transaction confirmation"],
    "high_value_vs_income": ["Bank statements covering the transaction dates",
                              "Source-of-funds proof (gift deed, loan agreement, prior savings, etc.)"],
    "gst_itr_gap": ["GST returns (GSTR-1, GSTR-3B) for the year",
                     "Reconciliation working between GST and ITR turnover"],
    "exemption_swap": ["Both original and revised ITR copies",
                        "Documentary proof for whichever exemption you intend to finally claim"],
    "stale_ais": ["Freshly downloaded AIS (same day as filing, if possible)"],
}


def build_workbook(analysis_result: dict, taxpayer_name: str = "", pan: str = "") -> Workbook:
    wb = Workbook()

    # --- Sheet 1: Mismatch Summary ---
    ws1 = wb.active
    ws1.title = "Mismatch Summary"
    ws1.append(["Notice Shield — Mismatch Summary"])
    ws1["A1"].font = Font(size=14, bold=True)
    ws1.append([f"Name: {taxpayer_name}", f"PAN: {pan}"])
    ws1.append([f"Overall verdict: {analysis_result['verdict'].upper()}"])
    ws1.append([])

    headers = ["Issue", "Severity", "Details", "What this means"]
    ws1.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws1.cell(row=5, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    row_idx = 6
    flags = analysis_result.get("flags", [])
    if not flags:
        ws1.append(["No mismatches found", "GREEN", "Your Form16 and AIS figures matched within tolerance.", "—"])
        for col in range(1, 5):
            ws1.cell(row=row_idx, column=col).fill = SEVERITY_FILL["green"]
    else:
        for flag in flags:
            label = BUCKET_LABELS.get(flag["bucket"], flag["bucket"])
            ws1.append([label, flag["severity"].upper(), flag["detail"], ""])
            for col in range(1, 5):
                ws1.cell(row=row_idx, column=col).fill = SEVERITY_FILL.get(flag["severity"], PatternFill())
                ws1.cell(row=row_idx, column=col).alignment = Alignment(wrap_text=True, vertical="top")
            row_idx += 1

    widths = [32, 12, 60, 30]
    for i, w in enumerate(widths, start=1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # --- Sheet 2: Document Checklist ---
    ws2 = wb.create_sheet("Document Checklist")
    ws2.append(["Notice Shield — Documents To Keep Ready"])
    ws2["A1"].font = Font(size=14, bold=True)
    ws2.append([])
    ws2.append(["Done?", "Document", "Related to issue"])
    for col in range(1, 4):
        cell = ws2.cell(row=3, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    row_idx = 4
    seen = set()
    for flag in flags:
        bucket = flag["bucket"]
        label = BUCKET_LABELS.get(bucket, bucket)
        for doc in CHECKLIST_BY_BUCKET.get(bucket, []):
            key = (doc, bucket)
            if key in seen:
                continue
            seen.add(key)
            ws2.append(["☐", doc, label])
            row_idx += 1

    if row_idx == 4:
        ws2.append(["—", "No specific documents needed — no issues were flagged.", ""])

    ws2.column_dimensions["A"].width = 8
    ws2.column_dimensions["B"].width = 55
    ws2.column_dimensions["C"].width = 35

    return wb


if __name__ == "__main__":
    from analysis import analyze

    form16 = {"gross_salary": {"value": 1250000}, "total_tds": {"value": 142000}}
    ais = {
        "salary_amount_reported": {"value": 1180000},
        "salary_tds_reported": {"value": 142000},
        "interest_income": {"value": 8200},
    }
    extra = {"original_itr_exemption": "HRA", "revised_itr_exemption": "10(14)"}
    result = analyze(form16, ais, extra)

    wb = build_workbook(result, taxpayer_name="Test User", pan="ABCDE1234A")
    wb.save("test_output.xlsx")
    print("Saved test_output.xlsx")
